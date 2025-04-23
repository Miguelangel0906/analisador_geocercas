import streamlit as st
import pandas as pd
from io import BytesIO
from fpdf import FPDF
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_excel(df: pd.DataFrame):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Geocercas')
        ws = writer.sheets['Geocercas']

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Estilizar encabezados
        for col_num, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border
            # Ajustar ancho de columna
            max_length = max(df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(col_num)].width = max_length

        # Estilizar datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = alignment
                cell.border = border

    buffer.seek(0)
    st.download_button(
        label="Exportar a Excel",
        data=buffer,
        file_name="geocercas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def export_pdf(df: pd.DataFrame):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', size=12)
    pdf.set_fill_color(79, 129, 189)  # color azul
    pdf.set_text_color(255, 255, 255)

    col_width = pdf.w / (len(df.columns) + 1)

    # Encabezado
    for col in df.columns:
        pdf.cell(col_width, 10, str(col), border=1, fill=True, align='C')
    pdf.ln()

    # Cuerpo
    pdf.set_font("Arial", size=10)
    pdf.set_text_color(0, 0, 0)
    for _, row in df.iterrows():
        for value in row:
            pdf.cell(col_width, 10, str(value), border=1, align='C')
        pdf.ln()

    buffer = BytesIO()
    pdf_bytes = pdf.output(dest='S').encode('latin-1')
    buffer.write(pdf_bytes)
    buffer.seek(0)

    st.download_button(
        label="Exportar a PDF",
        data=buffer,
        file_name="geocercas.pdf",
        mime="application/pdf"
    )
